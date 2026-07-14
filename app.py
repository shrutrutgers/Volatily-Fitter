import math
import os
import pickle
import tempfile
import time
from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from plotly.subplots import make_subplots
from openpyxl import Workbook

import volatility_fitting_daily as vf


SPX_DIV_YIELD = 0.0134
SPY_DIVS = [(0.25, 1.90), (0.50, 2.10), (0.75, 1.90), (1.00, 1.92)]
DATA_FETCH_VERSION = "six-tenors-no-1w-target-moneyness-repo-fallback-prior"
SNAPSHOT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "latest_run.pkl")
SNAPSHOT_MAX_AGE_SECONDS = 3600


def fallback_spy_div_yield(spot):
    return sum(div for _, div in SPY_DIVS) / spot if spot > 0 else 0.012


st.set_page_config(
    page_title="Volatility Surface Dashboard",
    page_icon="IV",
    layout="wide",
)


st.markdown(
    """
    <style>
    :root {
        --bg: #070b12;
        --panel: #0d121c;
        --panel-2: #141b27;
        --line: #263142;
        --text: #e7ecf5;
        --muted: #8b95a7;
        --blue: #68a8ff;
        --green: #36d16f;
        --cyan: #41d6c3;
        --amber: #d6a82e;
    }
    html, body, [data-testid="stAppViewContainer"], [data-testid="stApp"] {
        background: radial-gradient(circle at 20% 0%, #101827 0%, #070b12 36%, #05070b 100%);
        color: var(--text);
    }
    [data-testid="stHeader"] { background: rgba(7, 11, 18, 0); }
    [data-testid="stSidebar"] {
        background: #080d15;
        border-right: 1px solid var(--line);
    }
    [data-testid="stSidebar"] * { color: var(--text); }
    .block-container {
        max-width: 1180px;
        padding-top: 1.4rem;
        padding-bottom: 3rem;
    }
    h1, h2, h3 {
        letter-spacing: 0;
        color: var(--text);
    }
    p, label, span, div { color: inherit; }
    .app-hero {
        display: flex;
        align-items: center;
        gap: 14px;
        margin: 4px 0 18px 0;
    }
    .app-mark {
        width: 42px;
        height: 42px;
        border-radius: 10px;
        display: grid;
        place-items: center;
        color: white;
        font-weight: 800;
        background: linear-gradient(135deg, #5ea2ff, #8e64ff);
        box-shadow: 0 0 30px rgba(104, 168, 255, 0.22);
    }
    .app-title {
        font-size: 1.6rem;
        font-weight: 800;
        color: var(--text);
        line-height: 1.1;
    }
    .app-subtitle {
        margin-top: 4px;
        color: var(--muted);
        font-size: 0.95rem;
    }
    .pill {
        display: inline-block;
        margin-left: 10px;
        padding: 3px 8px;
        border-radius: 999px;
        border: 1px solid #244466;
        color: var(--blue);
        background: rgba(104, 168, 255, 0.10);
        font-size: 0.68rem;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        vertical-align: middle;
    }
    div[data-testid="stMetric"] {
        border: 1px solid var(--line);
        border-radius: 8px;
        padding: 12px 14px;
        background: linear-gradient(180deg, #151c28, #101722);
        box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.03);
    }
    div[data-testid="stMetricLabel"] p {
        color: var(--muted);
        text-transform: uppercase;
        font-size: 0.72rem;
        letter-spacing: 0.08em;
    }
    div[data-testid="stMetricValue"] {
        color: var(--text);
        font-weight: 800;
        font-size: 1.1rem;
    }
    .st-key-runtime-metric div[data-testid="stMetricValue"] {
        font-size: 0.9rem;
    }
    [data-testid="stVerticalBlockBorderWrapper"], [data-testid="stExpander"] {
        border-color: var(--line);
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background: var(--panel);
        border: 1px solid var(--line);
        border-radius: 10px;
        padding: 4px;
    }
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        color: var(--muted);
        padding: 8px 18px;
    }
    .stTabs [aria-selected="true"] {
        background: var(--panel-2);
        color: var(--text);
        box-shadow: 0 0 0 1px rgba(104, 168, 255, 0.15) inset;
    }
    .stButton > button, .stDownloadButton > button {
        border-radius: 8px;
        border: 1px solid #2d7c50;
        background: rgba(54, 209, 111, 0.14);
        color: #8ff0b1;
        font-weight: 700;
    }
    .stButton > button:hover, .stDownloadButton > button:hover {
        border-color: var(--green);
        background: rgba(54, 209, 111, 0.22);
        color: #c8ffdb;
    }
    [data-testid="stDataFrame"] {
        border: 1px solid var(--line);
        border-radius: 8px;
        overflow: hidden;
    }
    [data-testid="stMarkdownContainer"] p {
        color: var(--muted);
    }
    hr {
        border-color: var(--line);
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def get_secret(name):
    try:
        return st.secrets.get(name, "")
    except Exception:
        return ""


def load_excel_from_upload(uploaded_file):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name
    return vf.load_xlsx(tmp_path)


def sample_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "SPX"
    ws["B1"] = "Spot"
    ws["C1"] = 5667.56
    ws["A23"] = "SPY"
    ws["B23"] = "Spot"
    ws["C23"] = 566.76

    ws["A4"] = "Expiry"
    ws["B4"] = "2025-04-17"
    ws["C4"] = 0.0739726
    ws["D3"] = 0.045
    ws["E4"] = "Expiry"
    ws["F4"] = "2025-06-20"
    ws["G4"] = 0.2493151
    ws["H3"] = 0.045
    ws["I4"] = "Expiry"
    ws["J4"] = "2025-12-19"
    ws["K4"] = 0.7479452
    ws["L3"] = 0.0435

    ws["A5"] = "Strike"
    ws["B5"] = "C"
    ws["C5"] = "P"
    ws["E5"] = "Strike"
    ws["F5"] = "C"
    ws["G5"] = "P"
    ws["I5"] = "Strike"
    ws["J5"] = "C"
    ws["K5"] = "P"

    ws["A25"] = "Expiry"
    ws["B25"] = "2025-04-17"
    ws["C25"] = 0.0739726
    ws["D24"] = 0.045
    ws["E25"] = "Expiry"
    ws["F25"] = "2025-06-20"
    ws["G25"] = 0.2493151
    ws["H24"] = 0.045
    ws["I25"] = "Expiry"
    ws["J25"] = "2025-12-19"
    ws["K25"] = 0.7479452
    ws["L24"] = 0.0435

    ws["A26"] = "Strike"
    ws["B26"] = "C"
    ws["C26"] = "P"
    ws["E26"] = "Strike"
    ws["F26"] = "C"
    ws["G26"] = "P"
    ws["I26"] = "Strike"
    ws["J26"] = "C"
    ws["K26"] = "P"

    for row in range(6, 20):
        ws.cell(row, 1).value = 5000 + 100 * (row - 6)
        ws.cell(row, 2).value = "call_mid"
        ws.cell(row, 3).value = "put_mid"
        ws.cell(row, 5).value = 5000 + 100 * (row - 6)
        ws.cell(row, 6).value = "call_mid"
        ws.cell(row, 7).value = "put_mid"
        ws.cell(row, 9).value = 4500 + 200 * (row - 6)
        ws.cell(row, 10).value = "call_mid"
        ws.cell(row, 11).value = "put_mid"

    for row in range(27, 41):
        ws.cell(row, 1).value = 500 + 10 * (row - 27)
        ws.cell(row, 2).value = "call_mid"
        ws.cell(row, 3).value = "put_mid"
        ws.cell(row, 5).value = 500 + 10 * (row - 27)
        ws.cell(row, 6).value = "call_mid"
        ws.cell(row, 7).value = "put_mid"
        ws.cell(row, 9).value = 450 + 20 * (row - 27)
        ws.cell(row, 10).value = "call_mid"
        ws.cell(row, 11).value = "put_mid"

    rate_rows = [
        (2, 1, "m", 4.36), (3, 1.5, "m", 4.33), (5, 3, "m", 4.33),
        (7, 6, "m", 4.26), (8, 1, "y", 4.04), (9, 2, "y", 3.94),
    ]
    for row, mag, unit, rate in rate_rows:
        ws.cell(row, 16).value = mag
        ws.cell(row, 17).value = unit
        ws.cell(row, 18).value = rate

    ws["A43"] = "Template note: replace call_mid/put_mid placeholders with numeric mid prices."

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# 15-minute TTL: repeated runs reuse the snapshot instead of re-hitting
# yfinance/FRED, which rate-limit shared IPs on Streamlit Community Cloud.
@st.cache_data(ttl=900, show_spinner=False)
def load_latest_cached(fred_api_key, version):
    return vf.load_latest_data(fred_api_key)


def save_snapshot(data, source_label, results, fetched_at):
    try:
        with open(SNAPSHOT_PATH, "wb") as f:
            pickle.dump({
                "data": data,
                "source_label": source_label,
                "results": results,
                "saved_at": fetched_at,
            }, f)
    except Exception:
        pass


def load_snapshot():
    try:
        with open(SNAPSHOT_PATH, "rb") as f:
            return pickle.load(f)
    except Exception:
        return None


def apply_snapshot(snapshot, asset_choice):
    st.session_state["run_data"] = snapshot["data"]
    st.session_state["run_source_label"] = snapshot["source_label"]
    st.session_state["run_mode"] = "Fetch latest data"
    st.session_state["run_asset_choice"] = asset_choice
    st.session_state["run_asset_results"] = snapshot["results"]
    st.session_state["run_fetched_at"] = snapshot["saved_at"]


def analyze_asset(label, spot, divs, ydiv, quotes, curve, american, per_tenor, version):
    rate_curve = vf.build_curve(per_tenor, curve)
    div_list = divs

    if american:
        repo = vf.fit_repo_am(spot, rate_curve, div_list, quotes, ydiv=ydiv if ydiv > 0 else 0.012)
    else:
        repo = vf.fit_repo_eu(spot, rate_curve, div_list, quotes, ydiv=ydiv if ydiv > 0 else 0.012)

    ivs = vf.compute_ivs(spot, rate_curve, div_list, repo, quotes, american)
    coefs = vf.fit_surface(spot, rate_curve, div_list, repo, ivs)
    greeks = vf.compute_greeks(spot, rate_curve, div_list, repo, coefs, ivs, american, otm_only=True)

    repo_rows = []
    for t in sorted(repo):
        r = vf.interp_rate(rate_curve, t)
        s_eff = spot - vf.pv_divs(div_list, t, rate_curve)
        forward = s_eff * math.exp((r - repo[t]) * t)
        repo_rows.append({
            "asset": label,
            "tenor": t,
            "rate": r,
            "repo": repo[t],
            "forward": forward,
        })

    iv_rows = []
    greek_rows = []
    for t, rows in ivs.items():
        r = vf.interp_rate(rate_curve, t)
        forward = (spot - vf.pv_divs(div_list, t, rate_curve)) * math.exp((r - repo[t]) * t)

        def base_row(strike, is_call, iv):
            is_otm = (is_call and strike >= forward) or ((not is_call) and strike <= forward)
            return {
                "asset": label,
                "tenor": t,
                "strike": strike,
                "option": "Call" if is_call else "Put",
                "is_otm": is_otm,
                "iv": iv,
                "iv_percent": 100 * iv,
                "log_moneyness": math.log(strike / forward),
                "forward": forward,
            }

        for strike, is_call, iv in rows:
            iv_rows.append(base_row(strike, is_call, iv))
        for strike, is_call, iv, delta, skew_delta, gamma, skew_gamma in greeks.get(t, []):
            greek_rows.append({
                **base_row(strike, is_call, iv),
                "delta": delta,
                "skew_delta": skew_delta,
                "gamma": gamma,
                "skew_gamma": skew_gamma,
            })

    coef_rows = []
    for t, vals in sorted(coefs.items()):
        coef_rows.append({
            "asset": label,
            "tenor": t,
            "a": round(vals["a"], 6),
            "b": round(vals["b"], 6),
            "c": round(vals["c"], 6),
            "r2": round(vals["r2"], 4),
            "rmse": round(vals["rmse"], 4),
            "n": vals["n"],
        })

    return pd.DataFrame(repo_rows), pd.DataFrame(iv_rows), pd.DataFrame(coef_rows), pd.DataFrame(greek_rows)


def fitted_surface_figure(asset, iv_df, coef_df):
    fig = go.Figure()
    if iv_df.empty or coef_df.empty:
        return fig

    plot_df = iv_df[iv_df["is_otm"]].copy()
    if plot_df.empty:
        return fig

    tenor_ranges = {}
    for tenor in sorted(coef_df["tenor"].unique()):
        part = plot_df[plot_df["tenor"] == tenor]
        if not part.empty:
            tenor_ranges[tenor] = (float(part["log_moneyness"].min()), float(part["log_moneyness"].max()))

    if not tenor_ranges:
        return fig

    # Use a broad display range, but mask each tenor outside its own observed
    # range. That avoids quadratic extrapolation spikes without collapsing the
    # surface into a skinny ribbon when one tenor has a narrower strike range.
    x_min = max(float(plot_df["log_moneyness"].quantile(0.03)), -0.08)
    x_max = min(float(plot_df["log_moneyness"].quantile(0.97)), 0.06)
    x_grid = np.linspace(x_min, x_max, 60)
    tenors = sorted(coef_df["tenor"].unique())

    z_rows = []
    for tenor in tenors:
        row = coef_df.loc[coef_df["tenor"] == tenor].iloc[0]
        fitted = 100 * (row["a"] + row["b"] * x_grid + row["c"] * x_grid * x_grid)
        lo, hi = tenor_ranges[tenor]
        fitted = np.where((x_grid >= lo) & (x_grid <= hi), fitted, np.nan)
        z_rows.append(np.clip(fitted, 5, 35))

    fig.add_trace(go.Surface(
        x=x_grid,
        y=tenors,
        z=np.array(z_rows),
        colorscale="Viridis",
        opacity=0.86,
        name="Fitted surface",
        connectgaps=False,
        colorbar={"title": "IV %"},
    ))

    fig.add_trace(go.Scatter3d(
        x=plot_df["log_moneyness"],
        y=plot_df["tenor"],
        z=plot_df["iv_percent"],
        mode="markers",
        marker={"size": 4, "color": plot_df["iv_percent"], "colorscale": "Turbo"},
        name="OTM market IV points",
    ))

    fig.update_layout(
        title=f"{asset} Implied Volatility Surface",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font={"color": "#d8deea"},
        height=620,
        margin={"l": 0, "r": 0, "t": 50, "b": 0},
        scene={
            "bgcolor": "rgba(0,0,0,0)",
            "xaxis_title": "log(K / F)",
            "yaxis_title": "Time to expiry",
            "zaxis_title": "IV %",
            "xaxis": {"gridcolor": "#2a3444", "zerolinecolor": "#556174"},
            "yaxis": {"gridcolor": "#2a3444", "zerolinecolor": "#556174"},
            "zaxis": {"gridcolor": "#2a3444", "zerolinecolor": "#556174"},
        },
    )
    return fig


def tenor_label(tenor):
    days = tenor * 365
    if days < 75:
        return f"{days / 7:.1f} weeks"
    if days < 330:
        return f"{days / 30:.1f} months"
    return f"{tenor:.1f} year"


def smile_figure(asset, iv_df, coef_df):
    fig = go.Figure()
    palette = ["#68a8ff", "#f26b5b", "#16d0a5", "#a76cff", "#ffb347", "#28d7f5", "#f472b6", "#a3e16d"]
    for idx, tenor in enumerate(sorted(iv_df["tenor"].unique())):
        part = iv_df[(iv_df["tenor"] == tenor) & (iv_df["is_otm"])].sort_values("log_moneyness")
        if part.empty:
            continue
        color = palette[idx % len(palette)]
        label = tenor_label(tenor)
        fig.add_trace(go.Scatter(
            x=part["log_moneyness"],
            y=part["iv_percent"],
            mode="markers",
            name=f"Market {label}",
            text=part["option"] + " K=" + part["strike"].round(2).astype(str),
            marker={"color": color, "size": 7},
        ))

        coef = coef_df[coef_df["tenor"] == tenor]
        if not coef.empty:
            row = coef.iloc[0]
            x_grid = np.linspace(float(part["log_moneyness"].min()), float(part["log_moneyness"].max()), 100)
            fitted_iv = 100 * (row["a"] + row["b"] * x_grid + row["c"] * x_grid * x_grid)
            fit_label = f"Fit {label} (R²={row['r2']:.2f})"
            fig.add_trace(go.Scatter(
                x=x_grid,
                y=fitted_iv,
                mode="lines",
                name=fit_label,
                line={"width": 2, "color": color},
            ))

    fig.add_vline(
        x=0,
        line_dash="dash",
        line_color="#ff5b5b",
        annotation_text="ATM forward",
        annotation_position="top",
    )

    fig.update_layout(
        title=f"{asset} Volatility Smiles",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#0d121c",
        font={"color": "#d8deea"},
        height=420,
        margin={"l": 0, "r": 0, "t": 50, "b": 0},
        xaxis_title="log(K / F)",
        yaxis_title="IV %",
        legend_title="Tenor",
        xaxis={"gridcolor": "#2a3444", "zerolinecolor": "#556174"},
        yaxis={"gridcolor": "#2a3444", "zerolinecolor": "#556174"},
    )
    return fig


def greeks_figure(asset, greeks_df, coef_df, tenor):
    fig = make_subplots(
        rows=1, cols=2,
        subplot_titles=("OTM Delta vs Skew OTM Delta", "Gamma vs Skew-Gamma"),
        horizontal_spacing=0.08,
    )
    if greeks_df.empty:
        return fig

    part = greeks_df[(greeks_df["tenor"] == tenor) & (greeks_df["is_otm"])].sort_values("log_moneyness")
    if part.empty:
        return fig

    hover = part["option"] + " K=" + part["strike"].round(2).astype(str)
    is_put = part["option"] == "Put"
    # Put delta is negative (-1..0); call-equivalent = 1 + delta (== 1 - |delta|)
    # puts the put wing on the same 0..1 scale as calls so the smile reads as
    # one continuous curve instead of jumping across zero at the ATM boundary.
    delta_plot = part["delta"].where(~is_put, 1 + part["delta"])
    skew_delta_plot = part["skew_delta"].where(~is_put, 1 + part["skew_delta"])
    styles = [
        (delta_plot, skew_delta_plot, "#68a8ff", "#f26b5b", 1),
        (part["gamma"], part["skew_gamma"], "#68a8ff", "#f26b5b", 2),
    ]
    for plain, skew, plain_color, skew_color, col in styles:
        fig.add_trace(go.Scatter(
            x=part["log_moneyness"],
            y=plain,
            mode="lines+markers",
            name="Plain",
            legendgroup="plain",
            showlegend=(col == 1),
            text=hover,
            line={"width": 2, "color": plain_color},
            marker={"size": 7, "color": plain_color},
        ), row=1, col=col)
        fig.add_trace(go.Scatter(
            x=part["log_moneyness"],
            y=skew,
            mode="lines+markers",
            name="Skew",
            legendgroup="skew",
            showlegend=(col == 1),
            text=hover,
            line={"width": 2, "color": skew_color, "dash": "dash"},
            marker={"size": 7, "color": skew_color, "symbol": "diamond-open"},
        ), row=1, col=col)

    fig.update_layout(
        title=f"{asset} Greeks ({tenor_label(tenor)}, OTM options)",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#0d121c",
        font={"color": "#d8deea"},
        height=460,
        margin={"l": 0, "r": 0, "t": 70, "b": 0},
        legend_title="Convention",
    )
    for col, y_title in [(1, "OTM Delta"), (2, "Gamma")]:
        fig.update_xaxes(
            title_text="log(K / F)", gridcolor="#2a3444", zerolinecolor="#556174",
            row=1, col=col,
        )
        fig.update_yaxes(
            title_text=y_title, gridcolor="#2a3444", zerolinecolor="#556174",
            row=1, col=col,
        )
    return fig


def front_month_smile_figure(asset, iv_df, coef_df):
    fig = go.Figure()
    if iv_df.empty or coef_df.empty:
        return fig

    tenor = min(sorted(iv_df["tenor"].unique()), key=lambda t: abs(t - 30 / 365))
    part = iv_df[(iv_df["tenor"] == tenor) & (iv_df["is_otm"])].sort_values("log_moneyness")
    if part.empty:
        return fig
    fig.add_trace(go.Scatter(
        x=part["log_moneyness"],
        y=part["iv_percent"],
        mode="markers",
        name="Market IV",
        text=part["option"] + " K=" + part["strike"].round(2).astype(str),
        marker={"size": 7, "color": "#68a8ff"},
    ))

    coef = coef_df[coef_df["tenor"] == tenor]
    if not coef.empty:
        row = coef.iloc[0]
        x_grid = np.linspace(float(part["log_moneyness"].min()), float(part["log_moneyness"].max()), 100)
        fitted_iv = 100 * (row["a"] + row["b"] * x_grid + row["c"] * x_grid * x_grid)
        fig.add_trace(go.Scatter(
            x=x_grid,
            y=fitted_iv,
            mode="lines",
            name=f"Quadratic fit (R²={row['r2']:.2f})",
            line={"color": "#41d6c3", "width": 3},
        ))

    fig.add_vline(
        x=0,
        line_dash="dash",
        line_color="#ff5b5b",
        annotation_text="ATM forward",
        annotation_position="top",
    )
    fig.update_layout(
        title={"text": f"{asset} Front-Month Smile ({tenor_label(tenor)})", "x": 0.0, "y": 0.98},
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="#0d121c",
        font={"color": "#d8deea"},
        height=620,
        margin={"l": 0, "r": 0, "t": 70, "b": 0},
        xaxis_title="log(K / F)",
        yaxis_title="IV %",
        legend={
            "orientation": "v",
            "y": 0.98,
            "x": 0.98,
            "xanchor": "right",
            "yanchor": "top",
            "bgcolor": "rgba(13, 18, 28, 0.72)",
        },
        xaxis={"gridcolor": "#2a3444", "zerolinecolor": "#556174"},
        yaxis={"gridcolor": "#2a3444", "zerolinecolor": "#556174"},
    )
    return fig


def format_percent_table(df, cols):
    out = df.copy()
    for col in cols:
        if col in out:
            out[col] = 100 * out[col]
    return out


st.markdown(
    """
    <div class="app-hero">
      <div class="app-mark">IV</div>
      <div>
        <div class="app-title">Volatility Surface Dashboard <span class="pill">Live Market Data</span></div>
        <div class="app-subtitle">SPX/SPY implied volatility construction from Excel inputs or latest market data</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("Inputs")
    mode = st.radio("Data source", ["Fetch latest data", "Upload Excel file"])
    asset_choice = st.radio("Asset", ["SPX", "SPY", "Both"], horizontal=True)

    uploaded = None
    if mode == "Fetch latest data":
        st.caption("Rates are pulled from FRED using the app's configured secret.")
    else:
        uploaded = st.file_uploader("Upload OptionData.xlsx", type=["xlsx"])
        st.download_button(
            "Download Excel template",
            sample_workbook_bytes(),
            file_name="OptionData_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("Upload mode expects the same cell layout as the original project workbook.")

    run_clicked = st.button("Run analysis", type="primary", use_container_width=True)

st.markdown(
    "This dashboard fits a repo curve, computes implied volatilities, filters option data by moneyness, "
    "and visualizes the fitted quadratic surface."
)

# Streamlit is request-driven, so "refresh hourly" means: on every page load,
# check the age of what we have and refetch if it's older than an hour. New
# visitors get the last saved snapshot instantly; the first visit after startup
# (or after the snapshot goes stale) triggers a fresh fetch automatically.
auto_fetch = False
if not run_clicked and mode == "Fetch latest data":
    if st.session_state.get("run_mode") == "Fetch latest data" and "run_data" in st.session_state:
        if time.time() - st.session_state.get("run_fetched_at", 0) > SNAPSHOT_MAX_AGE_SECONDS:
            auto_fetch = True
    elif "run_data" not in st.session_state:
        snapshot = load_snapshot()
        if snapshot is not None and time.time() - snapshot["saved_at"] <= SNAPSHOT_MAX_AGE_SECONDS:
            apply_snapshot(snapshot, asset_choice)
        else:
            auto_fetch = True

if run_clicked or auto_fetch:
    try:
        with st.spinner("Loading market data..."):
            if mode == "Fetch latest data":
                fred_key = get_secret("FRED_API_KEY")
                if not fred_key:
                    st.error("Missing FRED_API_KEY in Streamlit secrets.")
                    st.stop()
                data = load_latest_cached(fred_key, DATA_FETCH_VERSION)
                if not data.get("spx_quotes") or not data.get("spy_quotes"):
                    # Partial fetch: don't let it stick in the 15-minute cache —
                    # clear the entry so the next run retries live — and backfill
                    # the missing asset from the last complete snapshot so users
                    # still see a surface (with a warning about its age).
                    load_latest_cached.clear()
                    fallback = load_snapshot()
                    if fallback is not None:
                        filled = vf.merge_partial_data(data, fallback["data"])
                        if filled:
                            data = dict(data)
                            data["backfilled_assets"] = filled
                            saved_dt = datetime.fromtimestamp(
                                fallback["saved_at"], ZoneInfo("America/New_York")
                            ).strftime("%Y-%m-%d %H:%M %Z")
                            data.setdefault("warnings", []).append(
                                f"{' and '.join(filled)} option chains are temporarily unavailable "
                                f"from yfinance — showing the last saved data from {saved_dt}. "
                                "Click Run analysis to retry."
                            )
                source_label = "Latest available yfinance option chains + FRED Treasury curve"
            else:
                if uploaded is None:
                    st.error("Upload an OptionData.xlsx file first.")
                    st.stop()
                data = load_excel_from_upload(uploaded)
                source_label = uploaded.name
    except Exception as exc:
        # A failed auto-refresh shouldn't blank the page if an older snapshot exists.
        snapshot = load_snapshot()
        if snapshot is not None and "run_data" not in st.session_state:
            st.warning(f"Live fetch failed ({exc}); showing the last saved snapshot instead.")
            apply_snapshot(snapshot, asset_choice)
        elif "run_data" in st.session_state:
            st.warning(f"Refresh failed ({exc}); continuing with previously loaded data.")
        else:
            st.error(str(exc))
            st.stop()
    else:
        # Cache across reruns: any widget touch (e.g. the Greeks expiry dropdown)
        # reruns this whole script and st.button reverts to False, so results must
        # survive outside the run_clicked branch or the page resets to the intro screen.
        st.session_state["run_data"] = data
        st.session_state["run_source_label"] = source_label
        st.session_state["run_mode"] = mode
        st.session_state["run_asset_choice"] = asset_choice
        st.session_state["run_asset_results"] = {}
        st.session_state["run_fetched_at"] = time.time()

if "run_data" not in st.session_state:
    st.info("Choose a data source in the sidebar and click Run analysis.")
    st.stop()

data = st.session_state["run_data"]
source_label = st.session_state["run_source_label"]
mode = st.session_state["run_mode"]
asset_choice = st.session_state["run_asset_choice"]

assets = []
if asset_choice in ("SPX", "Both"):
    if data.get("spx_quotes"):
        assets.append((
            "SPX", data["spx_spot"], [], data.get("spx_div_yield", SPX_DIV_YIELD),
            data["spx_quotes"], False, data["spx_rates"]
        ))
    else:
        st.warning("SPX option chains are temporarily unavailable from yfinance. Click Run analysis to retry.")
if asset_choice in ("SPY", "Both"):
    if data.get("spy_quotes"):
        assets.append((
            "SPY", data["spy_spot"], data.get("spy_divs", SPY_DIVS),
            data.get("spy_div_yield", fallback_spy_div_yield(data["spy_spot"])),
            data["spy_quotes"], True, data["spy_rates"]
        ))
    else:
        st.warning("SPY option chains are temporarily unavailable from yfinance. Click Run analysis to retry.")

if not assets:
    st.error("No usable option quotes are available for the selected asset.")
    st.stop()

st.subheader("Run Summary")
cols = st.columns(4)
cols[0].metric("Source", mode)
cols[1].metric("SPX spot", f"{data['spx_spot']:,.2f}")
cols[2].metric("SPY spot", f"{data['spy_spot']:,.2f}")
quote_time = data.get("quote_time")
if quote_time is not None:
    quote_ts = pd.Timestamp(quote_time)
    if quote_ts.tzinfo is None:
        quote_ts = quote_ts.tz_localize("UTC")
    quote_ts_et = quote_ts.tz_convert("America/New_York")
    cols[3].metric("Quotes as of (ET)", quote_ts_et.strftime("%Y-%m-%d %H:%M"))
    age_minutes = (pd.Timestamp.now(tz="America/New_York") - quote_ts_et).total_seconds() / 60
    if age_minutes > 30:
        st.caption(
            f"Latest option trade is {age_minutes / 60:.1f} hours old — the market is likely closed "
            "or quotes are delayed, so the surface reflects the last session."
        )
else:
    run_time_et = datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d %H:%M %Z")
    with cols[3]:
        with st.container(key="runtime-metric"):
            st.metric("Run time (ET)", run_time_et)
if mode == "Fetch latest data":
    fetched_at = st.session_state.get("run_fetched_at")
    fetched_note = ""
    if fetched_at:
        fetched_dt = datetime.fromtimestamp(fetched_at, ZoneInfo("America/New_York"))
        fetched_note = f" Data fetched {fetched_dt.strftime('%Y-%m-%d %H:%M %Z')};"
    st.caption(source_label + " —" + fetched_note + " auto-refreshes when more than an hour old.")
else:
    st.caption(source_label)
for warning in data.get("warnings", []):
    st.warning(warning)

asset_results = st.session_state["run_asset_results"]

for label, spot, divs, ydiv, quotes, american, rates in assets:
    if label not in asset_results:
        with st.spinner(f"Computing {label} surface..."):
            asset_results[label] = analyze_asset(
                label, spot, divs, ydiv, quotes, data["curve"], american, rates, DATA_FETCH_VERSION
            )
    repo_df, iv_df, coef_df, greeks_df = asset_results[label]

    st.divider()
    st.header(label)

    summary_cols = st.columns(4)
    summary_cols[0].metric("Spot", f"{spot:,.2f}")
    summary_cols[1].metric("Model", "American" if american else "European")
    summary_cols[2].metric("Tenors", len(repo_df))
    summary_cols[3].metric("IV points", len(iv_df))
    if label == "SPY":
        st.caption("Note: short-tenor SPY repo can be unstable because it is bootstrapped from limited/noisy ATM American option pairs. See the documentation Limitations section.")

    tab_surface, tab_smiles, tab_greeks, tab_tables = st.tabs(["Surface", "Smiles", "Greeks", "Tables"])

    with tab_surface:
        surface_col, skew_col = st.columns([2, 1])
        with surface_col:
            st.plotly_chart(fitted_surface_figure(label, iv_df, coef_df), use_container_width=True, key=f"{label}-surface")
        with skew_col:
            st.caption("Front month is displayed as the listed tenor closest to 30 calendar days.")
            st.plotly_chart(front_month_smile_figure(label, iv_df, coef_df), use_container_width=True, key=f"{label}-front-smile")

    with tab_smiles:
        st.caption("Tenor labels are shown in calendar weeks, months, or years. Fit legend entries include R² as a quick fit-quality indicator.")
        st.plotly_chart(smile_figure(label, iv_df, coef_df), use_container_width=True)

    with tab_greeks:
        st.caption(
            "OTM delta/gamma are Black-Scholes spot sensitivities of each out-of-the-money option, holding its "
            "fitted IV fixed. Skew OTM delta/gamma additionally account for the option sliding along the fitted "
            "smile as spot moves (sticky-moneyness convention). The delta chart plots puts as call-equivalent "
            "delta (1 + put delta) so the put and call wings read as one continuous curve; the table below still "
            "shows raw signed delta."
        )
        greek_tenors = sorted(greeks_df["tenor"].unique())
        if greek_tenors:
            nearest_tenor = min(greek_tenors)
            selected_tenor = st.selectbox(
                "Expiry",
                greek_tenors,
                index=greek_tenors.index(nearest_tenor),
                format_func=tenor_label,
                key=f"{label}-greeks-tenor",
            )
            st.plotly_chart(
                greeks_figure(label, greeks_df, coef_df, selected_tenor),
                use_container_width=True,
                key=f"{label}-greeks",
            )
            if american:
                st.info(
                    "A small kink at the ATM boundary (log-moneyness = 0) is expected, especially for longer "
                    "expiries. The curve switches from OTM puts to OTM calls at this point, and because American "
                    "options do not satisfy put-call parity exactly (due to early-exercise premium, particularly "
                    "with dividends), the two sides do not join perfectly. This effect grows with maturity, "
                    "making the kink more noticeable for longer-dated options."
                )
            else:
                st.caption(
                    "No kink is expected at the ATM boundary for SPX: European options satisfy put-call parity, "
                    "so the put and call wings join smoothly. Compare with SPY, where the American early-exercise "
                    "premium produces a visible kink."
                )
            st.markdown(f"**Greeks — {tenor_label(selected_tenor)} (OTM options)**")
            st.dataframe(
                greeks_df[greeks_df["tenor"] == selected_tenor],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "delta": st.column_config.NumberColumn("OTM Delta", format="%.4f"),
                    "skew_delta": st.column_config.NumberColumn("Skew OTM Delta", format="%.4f"),
                    "gamma": st.column_config.NumberColumn("Gamma", format="%.6f"),
                    "skew_gamma": st.column_config.NumberColumn("Skew Gamma", format="%.6f"),
                },
            )
        else:
            st.info("No greeks available for this asset.")

    with tab_tables:
        table_cols = st.columns([1.2, 1.2, 1.6])
        with table_cols[0]:
            st.markdown("**Repo Curve**")
            repo_show = format_percent_table(repo_df, ["rate", "repo"])
            st.dataframe(repo_show, use_container_width=True, hide_index=True)
        with table_cols[1]:
            st.markdown("**Surface Coefficients**")
            st.dataframe(coef_df, use_container_width=True, hide_index=True)
            st.caption("IV(x) = a + b x + c x^2, x = log(K/F). a: ATM level, b: skew, c: curvature.")
        with table_cols[2]:
            st.markdown("**Implied Vols**")
            st.dataframe(iv_df, use_container_width=True, hide_index=True)

    csv = iv_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        f"Download {label} IV data",
        csv,
        file_name=f"{label.lower()}_implied_vols.csv",
        mime="text/csv",
    )
    greeks_csv = greeks_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        f"Download {label} Greeks data",
        greeks_csv,
        file_name=f"{label.lower()}_greeks.csv",
        mime="text/csv",
    )

# Only persist complete, fully-fresh fetches: a partial or backfilled snapshot
# would overwrite the last good data with a mix of ages stamped as new.
if (mode == "Fetch latest data" and data.get("spx_quotes") and data.get("spy_quotes")
        and not data.get("backfilled_assets")):
    save_snapshot(
        data,
        source_label,
        st.session_state["run_asset_results"],
        st.session_state.get("run_fetched_at", time.time()),
    )
