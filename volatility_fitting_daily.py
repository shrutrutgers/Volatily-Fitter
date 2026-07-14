import math
import os
import sys
import time
from datetime import date, datetime

import numpy as np
from openpyxl import load_workbook


SQRT2PI = math.sqrt(2 * math.pi)
FRED_API_KEY = os.environ.get("FRED_API_KEY", "")
FRED_CACHE_TTL_SECONDS = 3600
_FRED_CURVE_CACHE = {"timestamp": 0.0, "curve": None}
Q_MIN = -0.02
Q_MAX = 0.12
# 0.1% relative spot bump: small enough for accurate central differences,
# large enough that American PDE interpolation noise doesn't dominate gamma.
GREEK_BUMP_REL = 0.001
SIG_FLOOR = 0.01


def npdf(x):
    return math.exp(-0.5 * x * x) / SQRT2PI


def ncdf(x):
    return 0.5 * (1 + math.erf(x / math.sqrt(2)))


def interp_rate(curve, t):
    pts = curve if isinstance(curve, list) and curve == sorted(curve) else sorted(curve)
    if t <= pts[0][0]:
        return pts[0][1]
    if t >= pts[-1][0]:
        return pts[-1][1]
    for i in range(1, len(pts)):
        t0, r0 = pts[i - 1]
        t1, r1 = pts[i]
        if t0 <= t <= t1:
            return r0 + (t - t0) / (t1 - t0) * (r1 - r0)
    return pts[-1][1]


def pv_divs(divs, T, curve):
    s = 0.0
    for ti, di in divs:
        if 0 < ti <= T:
            s += di * math.exp(-interp_rate(curve, ti) * ti)
    return s


def bs(S, K, T, r, q, sig, is_call):
    if T <= 0 or sig <= 0:
        return max(S - K, 0.0) if is_call else max(K - S, 0.0)
    sT = math.sqrt(T)
    d1 = (math.log(S / K) + (r - q + 0.5 * sig * sig) * T) / (sig * sT)
    d2 = d1 - sig * sT
    if is_call:
        return S * math.exp(-q * T) * ncdf(d1) - K * math.exp(-r * T) * ncdf(d2)
    return K * math.exp(-r * T) * ncdf(-d2) - S * math.exp(-q * T) * ncdf(-d1)


def bs_vega(S, K, T, r, q, sig):
    if T <= 0 or sig <= 0:
        return 0.0
    sT = math.sqrt(T)
    d1 = (math.log(S / K) + (r - q + 0.5 * sig * sig) * T) / (sig * sT)
    return S * math.exp(-q * T) * npdf(d1) * sT


def bs_delta(S, K, T, r, q, sig, is_call):
    if T <= 0 or sig <= 0:
        if is_call:
            return 1.0 if S > K else 0.0
        return -1.0 if S < K else 0.0
    sT = math.sqrt(T)
    d1 = (math.log(S / K) + (r - q + 0.5 * sig * sig) * T) / (sig * sT)
    if is_call:
        return math.exp(-q * T) * ncdf(d1)
    return math.exp(-q * T) * (ncdf(d1) - 1.0)


def bs_gamma(S, K, T, r, q, sig):
    if T <= 0 or sig <= 0:
        return 0.0
    sT = math.sqrt(T)
    d1 = (math.log(S / K) + (r - q + 0.5 * sig * sig) * T) / (sig * sT)
    return math.exp(-q * T) * npdf(d1) / (S * sig * sT)


# Crank-Nicolson PDE for American options on a log-spot grid.
# Uses the Brennan-Schwartz algorithm: a single back-substitution per timestep
# that enforces the early-exercise constraint inline. Works for puts (sweep
# upward from low S) and calls on dividend-paying stock (sweep downward).
def american_pde(S, K, T, r, q, sig, is_call, M=200, N=200, grid_center=None, grid_sig=None):
    if T <= 0:
        return max(S - K, 0.0) if is_call else max(K - S, 0.0)
    if sig <= 0:
        F = S * math.exp((r - q) * T)
        return math.exp(-r * T) * (max(F - K, 0.0) if is_call else max(K - F, 0.0))

    # grid_center/grid_sig let bumped-spot greek calls share one exact grid:
    # re-centering or re-scaling the grid per call shifts the payoff kink
    # relative to the nodes, and that discretization noise swamps h^2 in
    # finite-difference gammas.
    sT = (grid_sig if grid_sig else sig) * math.sqrt(T)
    x_mid = math.log(grid_center if grid_center else S)
    x_max = x_mid + 6 * sT
    x_min = x_mid - 6 * sT
    dx = (x_max - x_min) / M
    x = np.linspace(x_min, x_max, M + 1)
    Sg = np.exp(x)

    if is_call:
        V = np.maximum(Sg - K, 0.0)
    else:
        V = np.maximum(K - Sg, 0.0)
    payoff = V.copy()

    dt = T / N
    nu = r - q - 0.5 * sig * sig
    A_ = 0.25 * dt * (sig * sig / (dx * dx) - nu / dx)
    B_ = -0.5 * dt * (sig * sig / (dx * dx) + r)
    C_ = 0.25 * dt * (sig * sig / (dx * dx) + nu / dx)

    n_in = M - 1
    sub = -A_ * np.ones(n_in)
    diag = (1 - B_) * np.ones(n_in)
    sup = -C_ * np.ones(n_in)

    for n in range(N):
        rhs = A_ * V[:-2] + (1 + B_) * V[1:-1] + C_ * V[2:]

        tau = (n + 1) * dt
        if is_call:
            V_lo = 0.0
            V_hi = Sg[-1] * math.exp(-q * tau) - K * math.exp(-r * tau)
        else:
            V_lo = K * math.exp(-r * tau)
            V_hi = 0.0
        rhs[0] += A_ * V_lo
        rhs[-1] += C_ * V_hi

        # Brennan-Schwartz: forward elimination, then back-substitution that
        # clamps to the payoff. For a put the exercise region is at low S,
        # so we sweep from high S downward (standard tridiagonal direction).
        # For a call (only relevant with q>0) the exercise region is at high S,
        # so we reverse the system before solving.
        if is_call:
            # Reverse the tridiagonal so back-sweep starts from high S
            sub_r = sup[::-1].copy()
            sup_r = sub[::-1].copy()
            diag_r = diag[::-1].copy()
            rhs_r = rhs[::-1].copy()
            payoff_r = payoff[1:-1][::-1]
            sol_r = _bs_solve(sub_r, diag_r, sup_r, rhs_r, payoff_r)
            sol = sol_r[::-1]
        else:
            sol = _bs_solve(sub, diag, sup, rhs, payoff[1:-1])

        V[0] = V_lo
        V[-1] = V_hi
        V[1:-1] = sol

    # Local quadratic instead of linear interpolation: linear interp is flat in
    # its second derivative inside a cell, which would zero out finite-difference
    # gammas when bumped evaluations land within one grid cell. Evaluating at
    # the grid center (every non-greek call) returns the node value either way.
    xq = math.log(S)
    i = min(max(int(round((xq - x_min) / dx)), 1), M - 1)
    xm, xc, xp = x[i - 1], x[i], x[i + 1]
    lm = (xq - xc) * (xq - xp) / ((xm - xc) * (xm - xp))
    lc = (xq - xm) * (xq - xp) / ((xc - xm) * (xc - xp))
    lp = (xq - xm) * (xq - xc) / ((xp - xm) * (xp - xc))
    return float(lm * V[i - 1] + lc * V[i] + lp * V[i + 1])


def _bs_solve(sub, diag, sup, rhs, payoff):
    # Brennan-Schwartz: standard Thomas elimination, back-substitution applies
    # the constraint V >= payoff at each step.
    n = len(diag)
    cp = np.empty(n)
    dp = np.empty(n)
    cp[0] = sup[0] / diag[0]
    dp[0] = rhs[0] / diag[0]
    for i in range(1, n):
        m = diag[i] - sub[i] * cp[i - 1]
        cp[i] = sup[i] / m if i < n - 1 else 0.0
        dp[i] = (rhs[i] - sub[i] * dp[i - 1]) / m
    sol = np.empty(n)
    sol[-1] = max(dp[-1], payoff[-1])
    for i in range(n - 2, -1, -1):
        sol[i] = max(dp[i] - cp[i] * sol[i + 1], payoff[i])
    return sol


def iv_european(price, S, K, T, r, q, is_call):
    intrinsic = max((S * math.exp(-q * T) - K * math.exp(-r * T)) if is_call
                    else (K * math.exp(-r * T) - S * math.exp(-q * T)), 0.0)
    upper = (S * math.exp(-q * T)) if is_call else (K * math.exp(-r * T))
    if price <= intrinsic + 1e-12 or price >= upper - 1e-12:
        result = _bisect(lambda s: bs(S, K, T, r, q, s, is_call) - price, 1e-4, 5.0)
        return None if result < 0 else result

    sig = 0.2
    for _ in range(60):
        diff = bs(S, K, T, r, q, sig, is_call) - price
        if abs(diff) < 1e-8:
            return sig
        v = bs_vega(S, K, T, r, q, sig)
        if v < 1e-10:
            break
        sig -= diff / v
        if sig <= 1e-6 or sig > 5:
            break
    result = _bisect(lambda s: bs(S, K, T, r, q, s, is_call) - price, 1e-4, 5.0)
    return None if result < 0 else result


def iv_american(price, S, K, T, r, q, is_call, M=120, N=120):
    result = _bisect(lambda s: american_pde(S, K, T, r, q, s, is_call, M, N) - price,
                     1e-2, 1.5, tol=2e-4)
    return None if result < 0 else result


def _bisect(f, lo, hi, tol=1e-7, max_iter=80):
    flo, fhi = f(lo), f(hi)
    expand = 0
    while flo * fhi > 0 and expand < 6:
        hi *= 2
        fhi = f(hi)
        expand += 1
    if flo * fhi > 0:
        return lo if abs(flo) < abs(fhi) else hi
    for _ in range(max_iter):
        mid = 0.5 * (lo + hi)
        fm = f(mid)
        if abs(fm) < tol or 0.5 * (hi - lo) < tol:
            return mid
        if flo * fm < 0:
            hi, fhi = mid, fm
        else:
            lo, flo = mid, fm
    return 0.5 * (lo + hi)


def closest(strikes, target):
    return min(strikes, key=lambda k: abs(k - target))


def fit_repo_eu(S, curve, divs, quotes, ydiv=0.0):
    repo = {}
    q_fallback = ydiv if ydiv > 0 else 0.012
    q_prev = q_fallback
    for t in sorted(quotes.keys()):
        r = interp_rate(curve, t)
        S_eff = S - pv_divs(divs, t, curve)
        F_e = S_eff * math.exp((r - q_prev) * t)
        K = closest(sorted(quotes[t].keys()), F_e)
        C = quotes[t][K]["C"]
        P = quotes[t][K]["P"]
        F_imp = ((C - P) + K * math.exp(-r * t)) * math.exp(r * t)
        q_t = r - math.log(F_imp / S_eff) / t
        if q_t < Q_MIN or q_t > Q_MAX:
            q_t = q_prev
        repo[t] = q_t
        if Q_MIN < q_t < Q_MAX:
            q_prev = q_t
        else:
            q_prev = q_fallback
    return repo


def fit_repo_am(S, curve, divs, quotes, ydiv=0.0, M=80, N=80):
    repo = {}
    q_fallback = ydiv if ydiv > 0 else 0.012
    q_prev = q_fallback
    for t in sorted(quotes.keys()):
        r = interp_rate(curve, t)
        S_eff = S - pv_divs(divs, t, curve)
        F_e = S_eff * math.exp((r - q_prev) * t)
        K = closest(sorted(quotes[t].keys()), F_e)
        C = quotes[t][K]["C"]
        P = quotes[t][K]["P"]

        def diff(qq):
            call_iv = iv_american(C, S_eff, K, t, r, qq, True, M, N)
            put_iv = iv_american(P, S_eff, K, t, r, qq, False, M, N)
            if call_iv is None or put_iv is None:
                return None
            return call_iv - put_iv

        lo, hi = -0.02, 0.12
        flo, fhi = diff(lo), diff(hi)
        ex = 0
        while flo is not None and fhi is not None and flo * fhi > 0 and ex < 6:
            lo -= 0.05
            hi += 0.05
            flo, fhi = diff(lo), diff(hi)
            ex += 1
        if flo is None or fhi is None:
            q_t = q_prev
        elif flo * fhi > 0:
            q_t = lo if abs(flo) < abs(fhi) else hi
            if q_t < Q_MIN or q_t > Q_MAX:
                q_t = q_prev
        else:
            for _ in range(40):
                mid = 0.5 * (lo + hi)
                fm = diff(mid)
                if fm is None:
                    lo = hi = q_prev
                    break
                if abs(fm) < 1e-4 or 0.5 * (hi - lo) < 1e-6:
                    lo = hi = mid
                    break
                if flo * fm < 0:
                    hi, fhi = mid, fm
                else:
                    lo, flo = mid, fm
            q_t = 0.5 * (lo + hi)
            if q_t < Q_MIN or q_t > Q_MAX:
                q_t = q_prev
        repo[t] = q_t
        if Q_MIN < q_t < Q_MAX:
            q_prev = q_t
        else:
            q_prev = q_fallback
    return repo


def compute_ivs(S, curve, divs, repo, quotes, american, M=120, N=120):
    out = {}
    for t, sm in quotes.items():
        r = interp_rate(curve, t)
        q = repo[t]
        S_eff = S - pv_divs(divs, t, curve)
        rows = []
        for K in sorted(sm.keys()):
            for is_call, key in [(True, "C"), (False, "P")]:
                if key not in sm[K]:
                    continue
                price = sm[K][key]
                if american:
                    intrinsic = max(S_eff - K, 0.0) if is_call else max(K - S_eff, 0.0)
                    if price < intrinsic:
                        continue
                    iv = iv_american(price, S_eff, K, t, r, q, is_call, M, N)
                else:
                    iv = iv_european(price, S_eff, K, t, r, q, is_call)
                if iv is None or iv <= 0.0101 or iv > 2.0:
                    continue
                rows.append((K, is_call, iv))
        out[t] = rows
    return out


def fit_quad(xs, ys):
    A = np.column_stack([np.ones_like(xs), xs, xs * xs])
    coef, *_ = np.linalg.lstsq(A, ys, rcond=None)
    y_pred = coef[0] + coef[1] * xs + coef[2] * xs * xs
    ss_res = float(np.sum((ys - y_pred) ** 2))
    ss_tot = float(np.sum((ys - np.mean(ys)) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot > 1e-12 else float("nan")
    rmse = float(np.sqrt(ss_res / len(ys))) if len(ys) > 0 else float("nan")
    n_points = len(ys)
    return float(coef[0]), float(coef[1]), float(coef[2]), r2, rmse, n_points


def fit_surface(S, curve, divs, repo, ivs):
    out = {}
    for t, rows in ivs.items():
        r = interp_rate(curve, t)
        q = repo[t]
        S_eff = S - pv_divs(divs, t, curve)
        F = S_eff * math.exp((r - q) * t)
        xs, ys = [], []
        for K, is_call, iv in rows:
            otm = (is_call and K > F) or ((not is_call) and K < F)
            if otm and 0 < iv < 5:
                xs.append(math.log(K / F))
                ys.append(iv)
        a, b, c, r2, rmse, n_points = fit_quad(np.array(xs), np.array(ys))
        out[t] = {"a": a, "b": b, "c": c, "r2": r2, "rmse": rmse, "n": n_points}
    return out


def compute_greeks(S, curve, divs, repo, coefs, ivs, american, M=120, N=120):
    """Returns {tenor: [(K, is_call, iv, delta, skew_delta, gamma, skew_gamma), ...]}"""
    out = {}
    for t, rows in ivs.items():
        r = interp_rate(curve, t)
        q = repo[t]
        S_eff = S - pv_divs(divs, t, curve)
        cf = coefs.get(t)
        h = GREEK_BUMP_REL * S_eff
        res = []
        for K, is_call, sig in rows:
            if american:
                p_mid = american_pde(S_eff, K, t, r, q, sig, is_call, M, N, grid_center=S_eff)
                p_up = american_pde(S_eff + h, K, t, r, q, sig, is_call, M, N, grid_center=S_eff)
                p_dn = american_pde(S_eff - h, K, t, r, q, sig, is_call, M, N, grid_center=S_eff)
                delta = (p_up - p_dn) / (2 * h)
                gamma = (p_up - 2 * p_mid + p_dn) / (h * h)
            else:
                delta = bs_delta(S_eff, K, t, r, q, sig, is_call)
                gamma = bs_gamma(S_eff, K, t, r, q, sig)
                p_mid = bs(S_eff, K, t, r, q, sig, is_call)

            if cf is None:
                skew_delta, skew_gamma = delta, gamma
            else:
                # Reuse the already-fitted (a, b, c) at the bumped moneyness
                # rather than re-fitting per bump: the smile shape is assumed
                # sticky in log-moneyness, only the option's position on it moves.
                a, b, c = cf["a"], cf["b"], cf["c"]
                F = S_eff * math.exp((r - q) * t)
                x0 = math.log(K / F)
                sig_smile = max(a + b * x0 + c * x0 * x0, SIG_FLOOR)

                def smile_price(S_b):
                    F_b = S_b * math.exp((r - q) * t)
                    x_b = math.log(K / F_b)
                    sig_b = max(a + b * x_b + c * x_b * x_b, SIG_FLOOR)
                    if american:
                        return american_pde(S_b, K, t, r, q, sig_b, is_call, M, N,
                                            grid_center=S_eff, grid_sig=sig_smile)
                    return bs(S_b, K, t, r, q, sig_b, is_call)

                sp_up = smile_price(S_eff + h)
                sp_dn = smile_price(S_eff - h)
                # The gamma stencil mid must sit on the fitted smile too: the
                # quoted iv is off the fit by inversion tolerance + residual,
                # and that price gap divided by h^2 would swamp the true gamma.
                sp_mid = smile_price(S_eff)
                skew_delta = (sp_up - sp_dn) / (2 * h)
                skew_gamma = (sp_up - 2 * sp_mid + sp_dn) / (h * h)

            res.append((K, is_call, sig, delta, skew_delta, gamma, skew_gamma))
        out[t] = res
    return out


def load_xlsx(path="OptionData.xlsx"):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    spx_spot = float(ws["C1"].value)
    spy_spot = float(ws["C23"].value)

    curve = []
    for row in range(2, 16):
        mag = ws.cell(row, 16).value
        unit = ws.cell(row, 17).value
        rate = ws.cell(row, 18).value
        if not (isinstance(mag, (int, float)) and isinstance(rate, (int, float))):
            continue
        u = str(unit).strip().lower() if unit else ""
        if u.startswith("m"):
            tt = mag / 12
        elif u.startswith("y"):
            tt = float(mag)
        else:
            continue
        curve.append((tt, rate / 100))
    curve.sort()

    def read_block(start, end, tenor_row, rate_row):
        quotes, rates = {}, {}
        for k_col, c_col, p_col, t_col, r_col in [
            (1, 2, 3, 3, 4), (5, 6, 7, 7, 8), (9, 10, 11, 11, 12)
        ]:
            t = ws.cell(tenor_row, t_col).value
            if not isinstance(t, (int, float)):
                continue
            tk = round(float(t), 6)
            quotes.setdefault(tk, {})
            rt = ws.cell(rate_row, r_col).value
            if isinstance(rt, (int, float)):
                rates[tk] = float(rt)
            for r in range(start, end + 1):
                K = ws.cell(r, k_col).value
                C = ws.cell(r, c_col).value
                Pp = ws.cell(r, p_col).value
                if not isinstance(K, (int, float)):
                    continue
                e = {}
                if isinstance(C, (int, float)):
                    e["C"] = float(C)
                if isinstance(Pp, (int, float)):
                    e["P"] = float(Pp)
                if e:
                    quotes[tk][float(K)] = e
        return quotes, rates

    spx_q, spx_r = read_block(6, 19, 4, 3)
    spy_q, spy_r = read_block(27, 40, 25, 24)

    return {
        "spx_spot": spx_spot, "spy_spot": spy_spot, "curve": curve,
        "spx_quotes": spx_q, "spx_rates": spx_r,
        "spy_quotes": spy_q, "spy_rates": spy_r,
    }

def option_mid(row, warn_list=None):
    try:
        bid = float(row.get("bid", float("nan")))
        ask = float(row.get("ask", float("nan")))
        last = float(row.get("lastPrice", 0.0))
        volume = float(row.get("volume", 0.0))
        open_interest = float(row.get("openInterest", 0.0))
    except (TypeError, ValueError):
        return None

    if math.isnan(bid) or math.isnan(ask):
        if last > 0:
            if warn_list is not None:
                warn_list.append(row.name)
            return last
        return None

    if bid <= 0 or ask <= 0:
        return None

    spread = ask - bid
    mid = 0.5 * (bid + ask)

    c1 = mid > 0 and spread / mid < 0.3
    c2 = mid > 0.1
    c3 = ask > bid
    c4 = bid > 0
    c5 = volume > 0 or open_interest > 0

    if c1 and c2 and c3 and c4 and c5:
        return mid
    return None


def choose_expiries(expiry_strings, today):
    # Pick a balanced set of expiries so the surface has enough slices to look smooth.
    targets = [14 / 365, 30 / 365, 60 / 365, 90 / 365, 180 / 365, 365 / 365]
    dated = []
    for s in expiry_strings:
        d = datetime.strptime(s, "%Y-%m-%d").date()
        t = (d - today).days / 365
        if t > 0:
            dated.append((s, t))

    chosen = []
    for target in targets:
        if not dated:
            break
        s, t = min(dated, key=lambda x: abs(x[1] - target))
        if s not in [x[0] for x in chosen]:
            chosen.append((s, round(t, 6)))
    return chosen


def load_option_quotes(ticker, spot, selected_expiries):
    quotes, rates = {}, {}
    skipped = []
    for expiry, t in selected_expiries:
        chain = ticker.option_chain(expiry)
        strikes = sorted(set(chain.calls["strike"]).intersection(set(chain.puts["strike"])))

        # Center the strike sample around an approximate forward, not just spot.
        # Include a deliberate positive wing so the final repo-implied forward
        # still has OTM calls (K > F) after repo bootstrapping.
        forward_guess = spot * math.exp(0.04 * t)
        strikes = [k for k in strikes if 0.80 * spot <= k <= 1.25 * spot]
        target_xs = [-0.08, -0.06, -0.04, -0.03, -0.02, -0.01, 0.0,
                     0.01, 0.02, 0.03, 0.04, 0.06, 0.08, 0.10, 0.12]
        picked = []
        for x in target_xs:
            target_k = forward_guess * math.exp(x)
            nearest = min(strikes, key=lambda k: abs(k - target_k))
            if nearest not in picked:
                picked.append(nearest)
        strikes = sorted(picked)
        strikes = sorted(strikes)

        calls = chain.calls.set_index("strike")
        puts = chain.puts.set_index("strike")
        quotes[t] = {}
        stale_strikes = []

        for K in strikes:
            c = option_mid(calls.loc[K], stale_strikes)
            p = option_mid(puts.loc[K], stale_strikes)
            if c is None or p is None:
                continue
            quotes[t][float(K)] = {"C": c, "P": p}

        if stale_strikes:
            print(f"WARNING: {len(stale_strikes)} strikes used lastPrice fallback for expiry {expiry}: {stale_strikes}")

        if not quotes[t]:
            skipped.append(expiry)
            del quotes[t]

    if skipped:
        print(f"WARNING: skipped expiries with no usable option quotes: {skipped}")

    if not quotes:
        raise RuntimeError("No usable option quotes found for any selected expiry.")
    return quotes, rates


def fetch_fred_rate(series_id, api_key):
    import requests

    url = "https://api.stlouisfed.org/fred/series/observations"
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
        "sort_order": "desc",
        "limit": 10,
    }
    response = None
    for attempt in range(3):
        response = requests.get(url, params=params, timeout=20)
        if response.status_code != 429:
            break
        time.sleep(1.5 * (attempt + 1))

    if response.status_code >= 400:
        raise RuntimeError(f"FRED returned HTTP {response.status_code} for {series_id}.")

    observations = response.json().get("observations", [])

    for obs in observations:
        value = obs.get("value")
        if value and value != ".":
            return float(value) / 100
    return None


def load_fred_curve(api_key=FRED_API_KEY):
    if not api_key:
        raise RuntimeError("Missing FRED API key. Set FRED_API_KEY or enter it in the dashboard.")

    now = time.time()
    cached_curve = _FRED_CURVE_CACHE["curve"]
    if cached_curve is not None and now - _FRED_CURVE_CACHE["timestamp"] < FRED_CACHE_TTL_SECONDS:
        return list(cached_curve)

    # Treasury constant maturity series from FRED.
    # Rates come back in percent, so fetch_fred_rate converts them to decimals.
    series = [
        (1 / 12, "DGS1MO"),
        (3 / 12, "DGS3MO"),
        (6 / 12, "DGS6MO"),
        (1.0, "DGS1"),
        (2.0, "DGS2"),
        (5.0, "DGS5"),
    ]

    curve = []
    for t, series_id in series:
        rate = fetch_fred_rate(series_id, api_key)
        if rate is not None:
            curve.append((t, rate))

    if not curve:
        raise RuntimeError("Could not fetch Treasury rates from FRED.")
    curve.sort(key=lambda x: x[0])
    _FRED_CURVE_CACHE["timestamp"] = now
    _FRED_CURVE_CACHE["curve"] = list(curve)
    return curve


def load_latest_data(fred_api_key=FRED_API_KEY):
    try:
        import yfinance as yf
    except ImportError as exc:
        raise RuntimeError("yfinance is not installed. Install it or use Excel mode.") from exc

    fallback_spx_div_yield = 0.0134
    fallback_spy_divs = [(0.25, 1.90), (0.50, 2.10), (0.75, 1.90), (1.00, 1.92)]

    today = date.today()
    spx_ticker = yf.Ticker("^SPX")
    spy_ticker = yf.Ticker("SPY")

    try:
        spx_hist = spx_ticker.history(period="5d")
        spx_spot = float(spx_hist["Close"].dropna().iloc[-1])
    except Exception as exc:
        raise RuntimeError(f"Could not fetch SPX spot from yfinance: {exc}") from exc

    try:
        spy_hist = spy_ticker.history(period="5d")
        spy_spot = float(spy_hist["Close"].dropna().iloc[-1])
    except Exception as exc:
        raise RuntimeError(f"Could not fetch SPY spot from yfinance: {exc}") from exc

    try:
        spx_expiries = choose_expiries(spx_ticker.options, today)
    except Exception as exc:
        raise RuntimeError(f"Could not fetch SPX option expiries from yfinance: {exc}") from exc

    try:
        spy_expiries = choose_expiries(spy_ticker.options, today)
    except Exception as exc:
        raise RuntimeError(f"Could not fetch SPY option expiries from yfinance: {exc}") from exc

    if not spx_expiries or not spy_expiries:
        raise RuntimeError("Could not find enough option expiries from yfinance.")

    warnings = []
    try:
        spx_q, spx_r = load_option_quotes(spx_ticker, spx_spot, spx_expiries)
    except Exception as exc:
        spx_q, spx_r = {}, {}
        warnings.append(f"Could not fetch usable SPX option chains from yfinance: {exc}")

    try:
        spy_q, spy_r = load_option_quotes(spy_ticker, spy_spot, spy_expiries)
    except Exception as exc:
        spy_q, spy_r = {}, {}
        warnings.append(f"Could not fetch usable SPY option chains from yfinance: {exc}")

    if not spx_q and not spy_q:
        raise RuntimeError("Could not fetch usable SPX or SPY option chains from yfinance.")

    try:
        curve = sorted(load_fred_curve(fred_api_key), key=lambda x: x[0])
    except Exception as exc:
        raise RuntimeError(f"Could not fetch Treasury rates from FRED: {exc}") from exc

    spx_div_yield = fallback_spx_div_yield
    spy_divs = fallback_spy_divs
    spy_div_yield = sum(div for _, div in fallback_spy_divs) / spy_spot if spy_spot > 0 else 0.012
    try:
        raw_divs = spy_ticker.dividends.dropna().tail(4)
        recent_divs = [float(x) for x in raw_divs.tolist() if float(x) > 0]
        if len(recent_divs) == 4 and spy_spot > 0:
            spy_divs = [(0.25 * (i + 1), div) for i, div in enumerate(recent_divs)]
            spy_div_yield = sum(recent_divs) / spy_spot
            spx_div_yield = spy_div_yield
    except Exception as exc:
        print(f"WARNING: Could not fetch SPY dividends from yfinance; using fallback assumptions: {exc}")

    spx_r = {t: interp_rate(curve, t) for t in spx_q}
    spy_r = {t: interp_rate(curve, t) for t in spy_q}

    print("Fetched latest available data with yfinance.")
    print(f"  SPX spot: {spx_spot:.2f}")
    print(f"  SPY spot: {spy_spot:.2f}")
    print("  SPX expiries:", ", ".join(s for s, _ in spx_expiries))
    print("  SPY expiries:", ", ".join(s for s, _ in spy_expiries))
    print("  Treasury curve: fetched from FRED")
    print(f"  SPX dividend yield proxy: {spx_div_yield:.4%}")
    print(f"  SPY carry prior: {spy_div_yield:.4%}")
    print("  SPY dividend assumptions:", ", ".join(f"{t:.2f}y:${d:.2f}" for t, d in spy_divs))

    return {
        "spx_spot": spx_spot, "spy_spot": spy_spot, "curve": curve,
        "spx_quotes": spx_q, "spx_rates": spx_r,
        "spy_quotes": spy_q, "spy_rates": spy_r,
        "spx_div_yield": spx_div_yield, "spy_div_yield": spy_div_yield,
        "spy_divs": spy_divs,
        "warnings": warnings,
    }


def build_curve(per_tenor, fallback):
    if not per_tenor:
        return sorted(fallback, key=lambda x: x[0])
    c = sorted(per_tenor.items())
    max_t = max(t for t, _ in c)
    for t, r in fallback:
        if t > max_t:
            c.append((t, r))
    c.sort()
    return c


def run(label, S, divs, ydiv, quotes, curve, american, per_tenor=None):
    print("=" * 76)
    print(f" {label}   (American = {american})")
    print("=" * 76)

    rate_curve = build_curve(per_tenor, curve)
    div_list = divs

    if american:
        repo = fit_repo_am(S, rate_curve, div_list, quotes, ydiv=ydiv if ydiv > 0 else 0.012)
    else:
        repo = fit_repo_eu(S, rate_curve, div_list, quotes, ydiv=ydiv if ydiv > 0 else 0.012)

    print("\nRepo curve:")
    for t in sorted(repo.keys()):
        r = interp_rate(rate_curve, t)
        S_eff = S - pv_divs(div_list, t, rate_curve)
        F = S_eff * math.exp((r - repo[t]) * t)
        print(f"  t = {t:7.5f}   r = {100*r:7.4f}%   q = {100*repo[t]:7.4f}%"
              f"   F = {F:11.4f}")

    ivs = compute_ivs(S, rate_curve, div_list, repo, quotes, american)

    print("\nImplied vols:")
    for t in sorted(ivs.keys()):
        print(f"  t = {t}")
        for K, is_call, iv in sorted(ivs[t], key=lambda r: (r[0], r[1])):
            print(f"    K = {K:7.1f}  {'C' if is_call else 'P'}   IV = {100*iv:7.4f}%")

    coefs = fit_surface(S, rate_curve, div_list, repo, ivs)
    print("\nSurface coefficients  IV(x) = a + b*x + c*x^2,  x = ln(K/F):")
    for t in sorted(coefs.keys()):
        vals = coefs[t]
        print(f"  t = {t:7.5f}   a = {vals['a']: .6f}   b = {vals['b']: .6f}"
              f"   c = {vals['c']: .6f}   R2 = {vals['r2']: .4f}"
              f"   RMSE = {vals['rmse']: .4f}   n = {vals['n']}")

    return repo, ivs, coefs


def main():
    mode = sys.argv[1].lower() if len(sys.argv) >= 2 else "excel"

    if mode == "excel":
        path = sys.argv[2] if len(sys.argv) >= 3 else "OptionData.xlsx"
        print(f"Loading data from Excel file: {path}")
        data = load_xlsx(path)
    elif mode == "latest":
        try:
            data = load_latest_data()
        except RuntimeError as exc:
            print(exc)
            return None
    else:
        print("Usage:")
        print("  python volatility_fitting_daily.py excel OptionData.xlsx")
        print("  python volatility_fitting_daily.py latest")
        return None

    spx_div_yield = data.get("spx_div_yield", 0.0134)
    spy_divs = data.get("spy_divs", [(0.25, 1.90), (0.50, 2.10), (0.75, 1.90), (1.00, 1.92)])
    spy_div_yield = data.get("spy_div_yield", sum(div for _, div in spy_divs) / data["spy_spot"])

    spx = run("SPX", data["spx_spot"], [], spx_div_yield,
              data["spx_quotes"], data["curve"], False,
              per_tenor=data["spx_rates"])

    spy = run("SPY", data["spy_spot"], spy_divs, spy_div_yield,
              data["spy_quotes"], data["curve"], True,
              per_tenor=data["spy_rates"])

    return {"SPX": spx, "SPY": spy}


if __name__ == "__main__":
    main()
